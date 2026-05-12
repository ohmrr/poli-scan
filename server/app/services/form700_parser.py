import csv
from dataclasses import dataclass, field
from typing import List


@dataclass
class OfficialRecord:
    last_name: str
    first_name: str
    agency: str
    position: str
    holdings: List[str] = field(default_factory=list)

    @property
    def full_name(self) -> str:
        return f"{self.first_name} {self.last_name}".strip()


def normalize_headers(headers: list[str]) -> list[str]:
    return [h.replace("\n", " ").strip() if h else "" for h in headers]


def load_form700_csv(file_path: str) -> List[OfficialRecord]:
    officials: dict[str, OfficialRecord] = {}

    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        reader.fieldnames = normalize_headers(reader.fieldnames or [])

        for row in reader:
            last = row.get("Last Name", "").strip()
            first = row.get("First Name", "").strip()
            agency = row.get("Agency", "").strip()
            position = row.get("Position", "").strip()
            holding = row.get("NAME OF BUSINESS ENTITY", "").strip()

            if not last and not first:
                continue

            key = f"{last},{first},{agency}"

            if key not in officials:
                officials[key] = OfficialRecord(
                    last_name=last, first_name=first, agency=agency, position=position
                )

            if holding:
                officials[key].holdings.append(holding)

    return list(officials.values())


def load_form700_xlsx(file_path: str) -> List[OfficialRecord]:
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    officials: dict[str, OfficialRecord] = {}

    def find_header_row(ws, max_scan=5):
        for i, row in enumerate(ws.iter_rows(max_row=max_scan, values_only=True)):
            if row and str(row[0] or "").strip() == "Last Name":
                return i, list(normalize_headers([str(c) if c is not None else "" for c in row]))
        return None, None

    def get_or_create(last, first, agency, position):
        last = (last or "").strip()
        first = (first or "").strip()
        agency = (agency or "").strip()
        position = (position or "").strip()
        if not last and not first:
            return None
        key = f"{last},{first},{agency}"
        if key not in officials:
            officials[key] = OfficialRecord(
                last_name=last, first_name=first, agency=agency, position=position
            )
        return officials[key]

    def add_holding(record, value):
        if record and value:
            v = str(value).strip()
            if v:
                record.holdings.append(v)

    # Schedule A1 — stocks, REITs, investments
    if "Schedule A1" in wb.sheetnames:
        ws = wb["Schedule A1"]
        header_row_idx, headers = find_header_row(ws)
        if headers:
            for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
                if not row or not row[0]:
                    continue
                rec = get_or_create(row[0], row[1], row[3], row[4])
                entity_col = next((i for i, h in enumerate(headers) if "NAME OF BUSINESS ENTITY" in h), None)
                if entity_col is not None:
                    add_holding(rec, row[entity_col])

    # Schedule A-2 — business entities and trusts
    if "Schedule A-2" in wb.sheetnames:
        ws = wb["Schedule A-2"]
        header_row_idx, headers = find_header_row(ws)
        if headers:
            for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
                if not row or not row[0]:
                    continue
                rec = get_or_create(row[0], row[1], row[3], row[4])
                entity_col = next((i for i, h in enumerate(headers) if "NAME OF BUSINESS ENTITY OR TRUST" in h), None)
                if entity_col is not None:
                    add_holding(rec, row[entity_col])

    # Schedule B — real property
    if "Schedule B" in wb.sheetnames:
        ws = wb["Schedule B"]
        header_row_idx, headers = find_header_row(ws)
        if headers:
            addr_col = next((i for i, h in enumerate(headers) if "STREET ADDRESS" in h), None)
            city_col = next((i for i, h in enumerate(headers) if h == "CITY"), None)
            for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
                if not row or not row[0]:
                    continue
                rec = get_or_create(row[0], row[1], row[3], row[4])
                if addr_col is not None:
                    addr = str(row[addr_col] or "").strip()
                    city = str(row[city_col] or "").strip() if city_col is not None else ""
                    location = f"{addr}, {city}".strip(", ") if addr else city
                    add_holding(rec, location)

    # Schedule C — income sources (including spousal/partner income)
    if "Schedule C - Income Section" in wb.sheetnames:
        ws = wb["Schedule C - Income Section"]
        header_row_idx, headers = find_header_row(ws)
        if headers:
            for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
                if not row or not row[0]:
                    continue
                rec = get_or_create(row[0], row[1], row[3], row[4])
                source_col = next((i for i, h in enumerate(headers) if h == "NAME OF SOURCE"), None)
                if source_col is not None:
                    add_holding(rec, row[source_col])

    wb.close()
    return list(officials.values())
